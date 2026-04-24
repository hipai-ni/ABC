#!/usr/bin/env python3
# ==========     1. Imports & configuration        ==============
import re
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
CPU_LOG_DIR = ROOT / "sw/test/cpu/log"
CUDA_LOG_DIR = ROOT / "sw/test/cuda/log"
RESULTS = ROOT / "run/results"
BENCHMARK_XLSX = RESULTS / "benchmark_cycles.xlsx"
OUT_BY_DS = RESULTS / "by_datasize"
OUT_BY_CASE = RESULTS / "by_case"
OUT_BY_CASE_AVG_LAT = RESULTS / "by_case_avg_latency"
OUT_BY_CASE_TP = RESULTS / "by_case_throughput"
OUT_BY_DS_TP = RESULTS / "by_datasize_throughput"
DATA_SIZES = (1, 2, 4, 8, 16, 32, 64, 128, 256, 512)
# cpu / picorv32 + three GPU thread configs (see run_all_1.py GPU_THREAD_COUNTS -> sw/test/cuda/log/gpu_d{ds}_t{T}.log)
PLATFORMS = ("cpu", "picorv32", "gput16", "gput32", "gput64")
GPU_LOG_THREADS = {"gput16": 16, "gput32": 32, "gput64": 64}

BY_DATASIZE_YMAX = 2500


CASE_ROW2_HINTS = (
    "Simple loop",
    "Compute loop",
    "Simple loop\n+ simple branch",
    "For loop\n+ complex branch",
    "While loop\n+ complex branch",
    "Multiple layer loop",
    "Inductive loop",
    "Indirect memory access",
)
_CYCLE_RE = re.compile(r"test_case(\d+)\s+cycles\s*=\s*(\d+)")
_PMU_END_RE = re.compile(r"TB_PMU End; cnt =\s*(\d+)")


# ==========     2. Phase A: parse one log -> 8 cycle values (case0..case7)        ==============
def _cases_region(text: str) -> str:
    """Prefer the simulation region between PicoRV32/RTL markers when present."""
    a = text.find("All Cases 0~7 Start")
    b = text.find("All Cases Done!")
    if a != -1 and b != -1 and b > a:
        return text[a:b]
    return text


def parse_cycles_from_log(log_path: Path) -> list[int]:
    text = log_path.read_text(encoding="utf-8", errors="replace")
    found = [None] * 8
    for m in _CYCLE_RE.finditer(text):
        idx = int(m.group(1))
        if 0 <= idx <= 7:
            found[idx] = int(m.group(2))
    if all(v is not None for v in found):
        return [int(found[i]) for i in range(8)]

    # PicoRV32 (vvp): uses TB_PMU End; cnt = N instead of "test_caseN cycles ="
    region = _cases_region(text)
    pmu = _PMU_END_RE.findall(region)
    if len(pmu) >= 8:
        return [int(x) for x in pmu[:8]]

    missing = [i for i, v in enumerate(found) if v is None]
    raise ValueError(f"{log_path}: missing cycles for case(s) {missing} (no PMU fallback)")


def load_all_data() -> dict[str, dict[int, list[int]]]:
    """platform -> datasize -> [case0..case7]; GPU rows use cuda/log/gpu_d{ds}_t{threads}.log."""
    data: dict[str, dict[int, list[int]]] = {p: {} for p in PLATFORMS}
    for ds in DATA_SIZES:
        data["cpu"][ds] = parse_cycles_from_log(CPU_LOG_DIR / f"cpu_d{ds}.log")
        data["picorv32"][ds] = parse_cycles_from_log(CPU_LOG_DIR / f"picorv32_d{ds}.log")
        for plat, t in GPU_LOG_THREADS.items():
            data[plat][ds] = parse_cycles_from_log(CUDA_LOG_DIR / f"gpu_d{ds}_t{t}.log")
    return data


def _bar_offsets_width(n: int) -> tuple[tuple[float, ...], float]:
    """Centered bar offsets + bar width for n groups."""
    width = min(0.16, 0.82 / max(n, 1))
    half = (n - 1) / 2.0
    return tuple((i - half) * width for i in range(n)), width


# ==========     2b. Export preprocessed data -> Excel        ==============
def export_benchmarks_excel(data: dict[str, dict[int, list[int]]], out_path: Path) -> None:
    """One sheet: datasize, platform, case0..case7."""
    RESULTS.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "cycles"
    header = ["datasize", "platform"] + [f"case{i}" for i in range(8)]
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
    for ds in DATA_SIZES:
        for plat in PLATFORMS:
            ws.append([ds, plat] + data[plat][ds])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    for col in range(3, 11):
        ws.column_dimensions[get_column_letter(col)].width = 10
    wb.save(out_path)


def _by_datasize_xticklabels() -> list[str]:
    return [f"case{i}\n{CASE_ROW2_HINTS[i]}" for i in range(8)]


# ==========     3. Phase B: charts — fixed datasize (10 bar charts)        ==============
def plot_by_datasize(data: dict[str, dict[int, list[int]]]) -> None:
    OUT_BY_DS.mkdir(parents=True, exist_ok=True)
    x = np.arange(8)
    offsets, width = _bar_offsets_width(len(PLATFORMS))
    colors = ("#4477aa", "#228833", "#cc6677", "#dd8452", "#9467bd")
    for ds in DATA_SIZES:
        fig, ax = plt.subplots(figsize=(11, 6))
        for off, plat, lab, col in zip(offsets, PLATFORMS, PLATFORMS, colors):
            ax.bar(x + off, data[plat][ds], width, label=lab, color=col, clip_on=True)
        ax.set_ylim(0, BY_DATASIZE_YMAX)
        ax.set_xlabel("Case")
        ax.set_ylabel("Cycles")
        ax.set_title(f"Cycles by case (datasize = {ds})")
        ax.set_xticks(x)
        ax.set_xticklabels(_by_datasize_xticklabels(), fontsize=9)
        ax.legend()
        ax.grid(axis="y", linestyle="--", alpha=0.4)
        fig.tight_layout()
        fig.savefig(OUT_BY_DS / f"bar_datasize_{ds}.png", dpi=150, bbox_inches="tight")
        plt.close(fig)


# ==========     3b. Phase B: charts — same layout as by_datasize, y = throughput (datasize / cycles)        ==============
def plot_by_datasize_throughput(data: dict[str, dict[int, list[int]]]) -> None:
    """Grouped bar charts per datasize: x = case0..case7, y = datasize/cycles (same as by_case_throughput)."""
    OUT_BY_DS_TP.mkdir(parents=True, exist_ok=True)
    x = np.arange(8)
    offsets, width = _bar_offsets_width(len(PLATFORMS))
    colors = ("#4477aa", "#228833", "#cc6677", "#dd8452", "#9467bd")
    for ds in DATA_SIZES:
        fig, ax = plt.subplots(figsize=(11, 6))
        for off, plat, lab, col in zip(offsets, PLATFORMS, PLATFORMS, colors):
            ys = [float(ds) / data[plat][ds][ci] for ci in range(8)]
            ax.bar(x + off, ys, width, label=lab, color=col, clip_on=True)
        # ax.set_yscale("log")
        ax.set_xlabel("Case")
        ax.set_ylabel("Throughput (datasize / cycles)")
        ax.set_ylim(0,1.4)
        ax.set_title(f"Throughput by case (datasize = {ds})")
        ax.set_xticks(x)
        ax.set_xticklabels(_by_datasize_xticklabels(), fontsize=9)
        ax.legend()
        ax.grid(axis="y", linestyle="--", alpha=0.4)
        fig.tight_layout()
        fig.savefig(OUT_BY_DS_TP / f"bar_datasize_{ds}.png", dpi=150, bbox_inches="tight")
        plt.close(fig)


# ==========     4a. Phase B: charts — fixed case (8 line charts, linear x for datasize)        ==============
def plot_by_case(data: dict[str, dict[int, list[int]]]) -> None:
    OUT_BY_CASE.mkdir(parents=True, exist_ok=True)
    xs = list(DATA_SIZES)
    styles = ("-o", "-s", "-^", "-D", "-P")
    colors = ("#4477aa", "#228833", "#cc6677", "#dd8452", "#9467bd")
    for case_idx in range(8):
        fig, ax = plt.subplots(figsize=(11, 5))
        for plat, sty, lab, col in zip(PLATFORMS, styles, PLATFORMS, colors):
            ys = [data[plat][ds][case_idx] for ds in DATA_SIZES]
            ax.plot(xs, ys, sty, label=lab, color=col)
        ax.set_xlabel("Datasize")
        ax.set_ylabel("Cycles")
        ax.set_title(f"Cycles vs datasize (case {case_idx})")
        ax.set_xticks(xs)
        ax.legend()
        ax.grid(True, linestyle="--", alpha=0.4)
        fig.tight_layout()
        fig.savefig(OUT_BY_CASE / f"line_case_{case_idx}.png", dpi=150)
        plt.close(fig)


# ==========     4b. Phase B: charts — avg latency (cycles / datasize), fixed case (8 line charts)        ==============
def plot_by_case_avg_latency(data: dict[str, dict[int, list[int]]]) -> None:
    OUT_BY_CASE_AVG_LAT.mkdir(parents=True, exist_ok=True)
    xs = list(DATA_SIZES)
    styles = ("-o", "-s", "-^", "-D", "-P")
    colors = ("#4477aa", "#228833", "#cc6677", "#dd8452", "#9467bd")
    for case_idx in range(8):
        fig, ax = plt.subplots(figsize=(11, 5))
        for plat, sty, lab, col in zip(PLATFORMS, styles, PLATFORMS, colors):
            ys = [data[plat][ds][case_idx] / float(ds) for ds in DATA_SIZES]
            ax.plot(xs, ys, sty, label=lab, color=col)
        ax.set_xlabel("Datasize")
        ax.set_ylabel("Avg latency (cycles / datasize)")
        ax.set_title(f"Avg latency vs datasize (case {case_idx})")
        ax.set_xticks(xs)
        ax.legend()
        ax.grid(True, linestyle="--", alpha=0.4)
        fig.tight_layout()
        fig.savefig(OUT_BY_CASE_AVG_LAT / f"line_case_{case_idx}.png", dpi=150)
        plt.close(fig)


# ==========     4c. Phase B: charts — throughput (datasize / cycles), fixed case (8 line charts)        ==============
def plot_by_case_throughput(data: dict[str, dict[int, list[int]]]) -> None:
    OUT_BY_CASE_TP.mkdir(parents=True, exist_ok=True)
    xs = list(DATA_SIZES)
    styles = ("-o", "-s", "-^", "-D", "-P")
    colors = ("#4477aa", "#228833", "#cc6677", "#dd8452", "#9467bd")
    for case_idx in range(8):
        fig, ax = plt.subplots(figsize=(11, 5))
        for plat, sty, lab, col in zip(PLATFORMS, styles, PLATFORMS, colors):
            ys = [float(ds) / data[plat][ds][case_idx] for ds in DATA_SIZES]
            ax.plot(xs, ys, sty, label=lab, color=col)
        ax.set_xlabel("Datasize")
        ax.set_ylabel("Throughput (datasize / cycles)")
        ax.set_title(f"Throughput vs datasize (case {case_idx})")
        ax.set_xticks(xs)
        ax.legend()
        ax.grid(True, linestyle="--", alpha=0.4)
        fig.tight_layout()
        fig.savefig(OUT_BY_CASE_TP / f"line_case_{case_idx}.png", dpi=150)
        plt.close(fig)

# ==========     5. Entry        ==============
if __name__ == "__main__":
    bench = load_all_data()
    export_benchmarks_excel(bench, BENCHMARK_XLSX)
    plot_by_datasize(bench)
    plot_by_case(bench)
    plot_by_case_avg_latency(bench)
    plot_by_case_throughput(bench)
    plot_by_datasize_throughput(bench)
    print(f"Wrote Excel: {BENCHMARK_XLSX}")
    print(f"Wrote bar charts under: {OUT_BY_DS}")
    print(f"Wrote line charts under: {OUT_BY_CASE}")
    print(f"Wrote avg-latency line charts under: {OUT_BY_CASE_AVG_LAT}")
    print(f"Wrote throughput line charts under: {OUT_BY_CASE_TP}")
    print(f"Wrote throughput bar charts (by datasize) under: {OUT_BY_DS_TP}")
    print(f"All outputs under: {RESULTS}")
